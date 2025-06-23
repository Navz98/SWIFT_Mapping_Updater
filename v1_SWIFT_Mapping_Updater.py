import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Mapping Sheet Updater", layout="wide")
st.title("Mapping Sheet Updater")

def strip_all_string_columns(df):
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].astype(str).str.strip()
    return df

def build_path_column(df):
    path_stack = {}
    full_paths = []
    parent_child_keys = []

    for _, row in df.iterrows():
        level = row.get('Lvl')
        name = str(row.get('Name', '')).strip()
        tag = str(row.get('XML Tag', '')).strip()

        if pd.isna(level):
            full_paths.append(None)
            parent_child_keys.append(None)
            continue

        level = int(level)
        component = f"{tag}__{name}"
        path_stack[level] = component

        for lvl in list(path_stack.keys()):
            if lvl > level:
                del path_stack[lvl]

        sorted_levels = sorted(path_stack.keys())
        path = " > ".join([path_stack[lvl] for lvl in sorted_levels])
        full_paths.append(path)

        meaningful_tags = [path_stack[lvl] for lvl in sorted_levels if path_stack[lvl]]
        if len(meaningful_tags) >= 2:
            parent_child_key = " > ".join(meaningful_tags[-2:])
        else:
            parent_child_key = path
        parent_child_keys.append(parent_child_key)

    df['Hierarchy Path'] = full_paths
    df['Parent-Child Key'] = parent_child_keys
    return df

def process_excel(source_file, test_file):
    source_excel = pd.read_excel(source_file, sheet_name=None)
    test_excel = pd.read_excel(test_file, sheet_name=None)

    source_df = pd.concat(source_excel.values(), ignore_index=True)
    test_df = pd.concat(test_excel.values(), ignore_index=True)

    source_df = strip_all_string_columns(source_df)
    test_df = strip_all_string_columns(test_df)

    source_df = build_path_column(source_df)
    test_df = build_path_column(test_df)

    key_cols = ['Hierarchy Path', 'XML Tag']
    parent_child_key_col = 'Parent-Child Key'
    excluded_cols = key_cols + [parent_child_key_col, 'Level', 'Lvl']

    source_output_columns = [col for col in source_df.columns if col not in excluded_cols and not col.startswith('Unnamed')]
    test_output_columns = [col for col in test_df.columns if col not in excluded_cols and not col.startswith('Unnamed')]

    merge_columns = key_cols + source_output_columns

    source_clean = source_df[merge_columns].drop_duplicates(subset=key_cols)
    test_clean = test_df.copy()

    source_clean = source_clean.fillna("").astype(str).replace("nan", "")
    test_clean = test_clean.fillna("").astype(str).replace("nan", "")

    merged = pd.merge(test_clean, source_clean, on=key_cols, how='left', suffixes=('', '_source'))

    differences = []

    # 1. Exact match
    for col in test_output_columns:
        source_col = f"{col}_source"
        if source_col in merged.columns:
            diff_rows = merged[merged[col] != merged[source_col]]
            for _, row in diff_rows.iterrows():
                differences.append({
                    "Hierarchy Path": row["Hierarchy Path"],
                    "XML Tag": row["XML Tag"],
                    "Column": col,
                    "Test Value": row[col],
                    "Source Value": row.get(source_col, ""),
                    "Type": "Changed"
                })

    # 2. New and Missing rows
    source_keys = set(zip(source_clean['Hierarchy Path'], source_clean['XML Tag']))
    test_keys = set(zip(test_clean['Hierarchy Path'], test_clean['XML Tag']))

    for _, row in test_clean.iterrows():
        key = (row['Hierarchy Path'], row['XML Tag'])
        if key not in source_keys:
            for col in test_output_columns:
                differences.append({
                    "Hierarchy Path": row['Hierarchy Path'],
                    "XML Tag": row['XML Tag'],
                    "Column": col,
                    "Test Value": row.get(col, ""),
                    "Source Value": "",
                    "Type": "New in Test"
                })

    for _, row in source_clean.iterrows():
        key = (row['Hierarchy Path'], row['XML Tag'])
        if key not in test_keys:
            for col in test_output_columns:
                differences.append({
                    "Hierarchy Path": row['Hierarchy Path'],
                    "XML Tag": row['XML Tag'],
                    "Column": col,
                    "Test Value": "",
                    "Source Value": row.get(col, ""),
                    "Type": "Missing in Test"
                })

    # 3. Parent-Child fallback
    unmatched_keys = test_keys - source_keys
    test_unmatched = test_df[test_df.apply(lambda x: (x['Hierarchy Path'], x['XML Tag']) in unmatched_keys, axis=1)]

    fallback_merged = pd.merge(
        test_unmatched[[parent_child_key_col, 'XML Tag'] + test_output_columns],
        source_df[[parent_child_key_col, 'XML Tag'] + source_output_columns].drop_duplicates(),
        on=[parent_child_key_col, 'XML Tag'], how='inner', suffixes=('', '_source')
    )

    for col in test_output_columns:
        src_col = f"{col}_source"
        if src_col in fallback_merged.columns:
            fallback_diffs = fallback_merged[fallback_merged[col] != fallback_merged[src_col]]
            for _, row in fallback_diffs.iterrows():
                differences.append({
                    "Hierarchy Path": "",
                    "XML Tag": row["XML Tag"],
                    "Column": col,
                    "Test Value": row[col],
                    "Source Value": row[src_col],
                    "Type": "Changed (Fallback)"
                })

    # 4. Loose fallback — tag-only match, if tag is unique in both source and test
    source_counts = source_df['XML Tag'].value_counts()
    test_counts = test_df['XML Tag'].value_counts()

    unique_tags = [tag for tag in test_counts.index
                   if test_counts[tag] == 1 and source_counts.get(tag, 0) == 1]

    for tag in unique_tags:
        test_row = test_df[test_df['XML Tag'] == tag].iloc[0]
        source_row = source_df[source_df['XML Tag'] == tag].iloc[0]
        for col in test_output_columns:
            test_val = test_row.get(col, "").strip()
            source_val = source_row.get(col, "").strip()
            if test_val != source_val:
                differences.append({
                    "Hierarchy Path": "",
                    "XML Tag": tag,
                    "Column": col,
                    "Test Value": test_val,
                    "Source Value": source_val,
                    "Type": "Changed (Loose Match)"
                })

    differences_df = pd.DataFrame(differences)

    merged.drop(columns=[f"{col}_source" for col in source_output_columns if f"{col}_source" in merged.columns], inplace=True)
    merged = merged.astype(str).replace("nan", "")
    merged = merged.replace({r'_x000D_': ' ', r'\r': ' ', r'\n': ' '}, regex=True)

    stripped_source_export = source_df.drop(columns=['Hierarchy Path'], errors='ignore')
    stripped_source_export = stripped_source_export.replace("nan", "").replace({pd.NA: "", None: ""}).fillna("")
    stripped_source_export = stripped_source_export.replace({r'_x000D_': ' ', r'\r': ' ', r'\n': ' '}, regex=True)

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        stripped_source_export.to_excel(writer, sheet_name='Source', index=False)
        merged.to_excel(writer, sheet_name='New Mapping', index=False)
        if not differences_df.empty:
            differences_df.to_excel(writer, sheet_name='Differences', index=False)
        pd.DataFrame({
            "Color": ["Yellow", "Blue", "Red", "Green", "Light Green"],
            "Meaning": ["Changed", "New in Test", "Missing in Test", "Changed (Fallback)", "Changed (Loose Match)"]
        }).to_excel(writer, sheet_name="Legend", index=False)

    output.seek(0)
    wb = load_workbook(output)
    ws = wb["New Mapping"]
    header_map = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}

    fill_map = {
        "Changed": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "New in Test": PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid"),
        "Missing in Test": PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid"),
        "Changed (Fallback)": PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid"),
        "Changed (Loose Match)": PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    }

    for diff in differences:
        tag = diff["XML Tag"]
        col = diff["Column"]
        dtype = diff["Type"]
        if col not in header_map:
            continue
        for row in ws.iter_rows(min_row=2):
            if str(row[header_map["XML Tag"] - 1].value).strip() == tag:
                row[header_map[col] - 1].fill = fill_map.get(dtype, PatternFill())
                break

    if 'Hierarchy Path' in header_map:
        ws.delete_cols(header_map['Hierarchy Path'])

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)
    return final_output

# UI components
source_file = st.file_uploader("⬆️ Upload Source Mapping Excel File", type=[".xlsx"])
test_file = st.file_uploader("⬆️ Upload Target Excel File", type=[".xlsx"])

if source_file and test_file:
    if st.button("Do the trick ✨"):
        with st.spinner("🥁 Drum Rolls..."):
            result = process_excel(source_file, test_file)
            st.success("Ta Da! Click the below button to download.")
            st.download_button("📥 Download Updated Mapping Sheet", result, file_name="Updated_mapping_sheet.xlsx")

# Footer
st.markdown("""
    <style>
    .footer {
        position: fixed;
        bottom: 10px;
        width: 100%;
        text-align: center;
        color: grey;
        font-size: 0.9rem;
        font-family: 'Courier New', monospace;
    }
    </style>
    <div class="footer">
         🧘‍♂️ Designed By Naveen 🧘‍♂️
    </div>
""", unsafe_allow_html=True)
